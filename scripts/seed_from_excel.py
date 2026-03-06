"""
Seed PostgreSQL from Caspio Excel exports.

Reads the 4 Caspio-exported XLSX files and inserts all data
into the cases, dockets, documents, and secondary_sources tables.

Usage (from project root):
    python scripts/seed_from_excel.py

Requires: openpyxl, psycopg2-binary
Connects to PostgreSQL on localhost:5432 (docker-compose exposed port).
"""

import os
import sys
from datetime import date, datetime

import openpyxl
import psycopg2
from psycopg2.extras import execute_values

# ── Configuration ────────────────────────────────────────────────────
DB_DSN = os.environ.get(
    "DATABASE_URL_SYNC",
    "host=localhost port=5432 dbname=dail user=dail_user password=dail_secret_changeme",
)
# If we get a SQLAlchemy-style URL, convert to psycopg2 DSN
if DB_DSN.startswith("postgresql"):
    # e.g. postgresql+asyncpg://user:pass@host:port/db → host=… port=… …
    from urllib.parse import urlparse
    u = urlparse(DB_DSN.replace("+asyncpg", ""))
    DB_DSN = f"host={u.hostname} port={u.port or 5432} dbname={u.path.lstrip('/')} user={u.username} password={u.password}"

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

EXCEL_FILES = {
    "cases": os.path.join(BASE_DIR, "Case_Table_2026-Feb-21_1952.xlsx"),
    "dockets": os.path.join(BASE_DIR, "Docket_Table_2026-Feb-21_2003.xlsx"),
    "documents": os.path.join(BASE_DIR, "Document_Table_2026-Feb-21_2002.xlsx"),
    "secondary_sources": os.path.join(
        BASE_DIR, "Secondary_Source_Coverage_Table_2026-Feb-21_2058.xlsx"
    ),
}

# ── Column Mappings: Excel header → DB column name ───────────────────
# Keys that don't appear here are silently skipped.

CASE_MAP = {
    "Case_snug": "case_slug",
    "Record_Number": "record_number",
    "Caption": "caption",
    "Brief_Description": "brief_description",
    "Area_of_Application_List": "area_of_application",
    "Area_of_Application_Text": "area_of_application_text",
    "Issue_Text": "issue_text",
    "Issue_List": "issue_list",
    "Cause_of_Action_List": "cause_of_action_list",
    "Cause_of_Action_Text": "cause_of_action_text",
    "Name_of_Algorithm_List": "name_of_algorithm_list",
    "Name_of_Algorithm_Text": "name_of_algorithm_text",
    "Class_Action_list": "class_action_list",
    "Class_Action": "class_action",
    "Organizations_involved": "organizations_involved",
    "Jurisdiction_Filed": "jurisdiction_filed",
    "Date_Action_Filed": "date_action_filed",
    "Current_Jurisdiction": "current_jurisdiction",
    "Jurisdiction_Type": "jurisdiction_type",
    "Jurisdiction_Name": "jurisdiction_name",
    "Jurisdiction_Type_Text": "jurisdiction_type_text",
    "Published_Opinions": "published_opinions",
    "Published_Opinions_binary": "published_opinions_binary",
    "Status_Disposition": "status_disposition",
    "Date_Added": "date_added",
    "Last_Update": "last_update",
    "Progress_Notes": "progress_notes",
    "Researcher": "researcher",
    "Summary_of_Significance": "summary_of_significance",
    "Summary_Facts_Activity_to_Date": "summary_facts_activity",
    "Most_Recent_Activity": "most_recent_activity",
    "Most_Recent_Activity_Date": "most_recent_activity_date",
    "Keyword": "keyword",
}

DOCKET_MAP = {
    "Case_Number": "case_number",
    "court": "court",
    "number": "number",
    "link": "link",
}

DOCUMENT_MAP = {
    "Case_Number": "case_number",
    "court": "court",
    "date": "date",
    "link": "link",
    "cite_or_reference": "cite_or_reference",
    "document": "document",
}

SECONDARY_SOURCE_MAP = {
    "Case_Number": "case_number",
    "Secondary_Source_Link": "secondary_source_link",
    "Secondary_Source_Title": "secondary_source_title",
}


# ── Helper Functions ─────────────────────────────────────────────────


def read_excel_data(filepath: str, column_map: dict) -> list[dict]:
    """Read an Excel file and return a list of row dicts mapped to DB columns."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]  # First sheet is the data sheet

    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return []

    # First row is the header
    headers = [str(h) if h else None for h in rows[0]]

    # Build index: position → db_column_name
    col_indices = {}
    for i, header in enumerate(headers):
        if header and header in column_map:
            col_indices[i] = column_map[header]

    result = []
    for row in rows[1:]:
        record = {}
        for pos, db_col in col_indices.items():
            val = row[pos] if pos < len(row) else None
            record[db_col] = clean_value(val, db_col)
        result.append(record)

    return result


def clean_value(val, col_name: str):
    """Normalize a cell value for PostgreSQL insertion."""
    if val is None:
        return None

    # Convert datetime → date for date-only columns
    if isinstance(val, datetime):
        if col_name in (
            "date_action_filed",
            "most_recent_activity_date",
            "date",
        ):
            return val.date()
        return val

    # Boolean handling
    if col_name == "published_opinions_binary":
        if isinstance(val, (int, float)):
            return bool(int(val))
        if isinstance(val, str):
            return val.lower() in ("1", "true", "yes", "y")
        return False

    # Integer handling
    if col_name in ("record_number", "case_number"):
        if isinstance(val, (int, float)):
            return int(val)
        try:
            return int(val)
        except (ValueError, TypeError):
            return None

    # Strip whitespace from strings
    if isinstance(val, str):
        val = val.strip()
        return val if val else None

    return val


def insert_cases(cur, records: list[dict]) -> int:
    """Insert case records. Skip rows without a record_number."""
    if not records:
        return 0

    # Filter out rows missing critical fields
    valid = [r for r in records if r.get("record_number") is not None and r.get("caption")]

    if not valid:
        return 0

    columns = sorted(valid[0].keys())
    cols_sql = ", ".join(columns)
    template = "(" + ", ".join(f"%({c})s" for c in columns) + ")"

    # Use ON CONFLICT to skip duplicates on record_number
    sql = f"""
        INSERT INTO cases ({cols_sql})
        VALUES {template}
        ON CONFLICT (record_number) DO NOTHING
    """

    count = 0
    for rec in valid:
        try:
            cur.execute(sql.replace(template, template), rec)
            count += cur.rowcount
        except Exception as e:
            print(f"  WARN: Skipping case record_number={rec.get('record_number')}: {e}")
            cur.connection.rollback()
            continue

    return count


def insert_child_table(cur, table: str, records: list[dict], fk_col: str = "case_number") -> int:
    """Insert records into a child table, skipping rows whose FK doesn't exist in cases."""
    if not records:
        return 0

    # Filter rows that have a valid case_number
    valid = [r for r in records if r.get(fk_col) is not None]
    if not valid:
        return 0

    columns = sorted(valid[0].keys())
    cols_sql = ", ".join(columns)
    placeholders = ", ".join(f"%({c})s" for c in columns)

    sql = f"INSERT INTO {table} ({cols_sql}) VALUES ({placeholders})"

    count = 0
    for rec in valid:
        try:
            cur.execute(sql, rec)
            count += cur.rowcount
        except Exception as e:
            print(f"  WARN: Skipping {table} row case_number={rec.get(fk_col)}: {e}")
            cur.connection.rollback()
            continue

    return count


def main():
    # Verify all Excel files exist
    for name, path in EXCEL_FILES.items():
        if not os.path.exists(path):
            print(f"ERROR: Missing Excel file for {name}: {path}")
            sys.exit(1)

    print("Connecting to PostgreSQL...")
    conn = psycopg2.connect(DB_DSN)
    conn.autocommit = False
    cur = conn.cursor()

    try:
        # ── Step 1: Clear existing data (child tables first due to FK) ──
        print("Clearing existing data...")
        for table in ("secondary_sources", "documents", "dockets", "cases"):
            cur.execute(f"DELETE FROM {table}")
            print(f"  Cleared {table}")
        conn.commit()

        # Reset sequences
        for table in ("cases", "dockets", "documents", "secondary_sources"):
            cur.execute(f"SELECT setval(pg_get_serial_sequence('{table}', 'id'), 1, false)")
        conn.commit()

        # ── Step 2: Load Cases ──────────────────────────────────────────
        print("\nLoading cases from Excel...")
        case_records = read_excel_data(EXCEL_FILES["cases"], CASE_MAP)
        print(f"  Read {len(case_records)} rows from Excel")

        count = insert_cases(cur, case_records)
        conn.commit()
        print(f"  Inserted {count} cases")

        # Verify which record_numbers are in the DB (for FK validation)
        cur.execute("SELECT record_number FROM cases")
        valid_rns = {row[0] for row in cur.fetchall()}
        print(f"  {len(valid_rns)} valid record_numbers in DB")

        # ── Step 3: Load Dockets ────────────────────────────────────────
        print("\nLoading dockets from Excel...")
        docket_records = read_excel_data(EXCEL_FILES["dockets"], DOCKET_MAP)
        # Filter to only FK-valid rows
        docket_records = [r for r in docket_records if r.get("case_number") in valid_rns]
        print(f"  {len(docket_records)} rows with valid case references")

        count = insert_child_table(cur, "dockets", docket_records)
        conn.commit()
        print(f"  Inserted {count} dockets")

        # ── Step 4: Load Documents ──────────────────────────────────────
        print("\nLoading documents from Excel...")
        doc_records = read_excel_data(EXCEL_FILES["documents"], DOCUMENT_MAP)
        doc_records = [r for r in doc_records if r.get("case_number") in valid_rns]
        print(f"  {len(doc_records)} rows with valid case references")

        count = insert_child_table(cur, "documents", doc_records)
        conn.commit()
        print(f"  Inserted {count} documents")

        # ── Step 5: Load Secondary Sources ──────────────────────────────
        print("\nLoading secondary sources from Excel...")
        ss_records = read_excel_data(EXCEL_FILES["secondary_sources"], SECONDARY_SOURCE_MAP)
        ss_records = [r for r in ss_records if r.get("case_number") in valid_rns]
        print(f"  {len(ss_records)} rows with valid case references")

        count = insert_child_table(cur, "secondary_sources", ss_records)
        conn.commit()
        print(f"  Inserted {count} secondary sources")

        # ── Summary ─────────────────────────────────────────────────────
        print("\n" + "=" * 50)
        print("SEED COMPLETE — Final counts:")
        for table in ("cases", "dockets", "documents", "secondary_sources"):
            cur.execute(f"SELECT COUNT(*) FROM {table}")
            n = cur.fetchone()[0]
            print(f"  {table}: {n} rows")
        print("=" * 50)

    except Exception as e:
        conn.rollback()
        print(f"\nERROR: {e}")
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    main()
